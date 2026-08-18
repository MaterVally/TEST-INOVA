"""
Excel preprocessing pipeline.

Enterprise Compliance Intelligence Platform

Purpose
-------
Process Excel workbooks into the same format expected by MMKGBuilder.

Output
------
texts  : List[str]
images : List[dict]

Images are currently empty because embedded image extraction
is reserved for V2.

Compatible with:
- .xlsx
- .xlsm
"""

from __future__ import annotations

from openpyxl import load_workbook

from ..utils.base import logger


class ExcelChunking:

    def __init__(
        self,
        excel_path: str,
        working_dir: str,
    ):

        self.excel_path = excel_path
        self.working_dir = working_dir

    # ---------------------------------------------------------
    # Public Entry
    # ---------------------------------------------------------

    async def process(
        self
    ) -> tuple[list[str], list[dict]]:

        logger.info("📊 Processing Excel workbook...")

        texts = self._extract_text()

        images = []

        logger.info(
            f"✅ Excel Parsed "
            f"({len(texts)} rows extracted)"
        )

        return texts, images

    # ---------------------------------------------------------
    # Extract workbook contents
    # ---------------------------------------------------------

    def _extract_text(self) -> list[str]:

        workbook = load_workbook(
            self.excel_path,
            data_only=True
        )

        extracted = []

        for worksheet in workbook.worksheets:

            extracted.append(
                f"=== SHEET: {worksheet.title} ==="
            )

            # Read header row first (first non-empty row)
            headers: list[str] = []
            rows_iter = worksheet.iter_rows(values_only=True)

            for row in rows_iter:
                candidate = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if candidate:
                    headers = candidate
                    break

            if not headers:
                continue

            # Emit a header summary so the LLM knows the schema
            extracted.append(f"This sheet contains columns: {', '.join(headers)}.")

            # Convert each data row into a natural-language sentence
            for row in rows_iter:
                values = [str(v).strip() if v is not None else "" for v in row]

                # Skip completely empty rows
                if not any(values):
                    continue

                # Build "Column: Value" pairs, skip empty cells
                pairs = [
                    f"{h}: {v}"
                    for h, v in zip(headers, values)
                    if v
                ]

                if pairs:
                    extracted.append(". ".join(pairs) + ".")

        workbook.close()

        return extracted
